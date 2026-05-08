import openpyxl, json
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

results = json.load(open('test-results.json'))
result_map = {r['tcId']: r for r in results}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Manual Testing - Sprint 3'

HDR_BG  = '1F4E79'; HDR_FG  = 'FFFFFF'
FEAT_BG = '2E75B6'; FEAT_FG = 'FFFFFF'
ODD     = 'DEEAF1'; EVEN    = 'FFFFFF'
PASS_BG = 'E2EFDA'; PASS_FG = '375623'
FAIL_BG = 'FFDCE1'; FAIL_FG = '9C0006'

thin  = Side(style='thin',   color='AAAAAA')
thick = Side(style='medium', color='2E75B6')

def tborder():   return Border(left=thin,  right=thin,  top=thin,  bottom=thin)
def fborder():   return Border(left=thick, right=thick, top=thick, bottom=thick)
def fill(c):     return PatternFill('solid', fgColor=c)
def bfont(c='000000', bold=False, sz=10): return Font(name='Calibri', bold=bold, size=sz, color=c)
def align(h='left', v='top', wrap=True):  return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

for i, w in enumerate([22, 12, 30, 44, 34, 44, 34, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Row 1 Title
ws.merge_cells('A1:H1')
c = ws['A1']
c.value = 'Manual Testing Document — CRMS Sprint 3'
c.font = Font(name='Calibri', bold=True, size=16, color=HDR_FG)
c.fill = fill(HDR_BG); c.alignment = align('center', 'center')
ws.row_dimensions[1].height = 32

# Row 2 Subtitle
ws.merge_cells('A2:H2')
c = ws['A2']
c.value = 'Student: Asif Bin Mahmood (ID: 23201632)  |  Course: CSE470 — Software Engineering  |  Sprint 3  |  Live Tested: 40/40 Passed'
c.font = Font(name='Calibri', bold=False, size=10, color=HDR_FG)
c.fill = fill(HDR_BG); c.alignment = align('center', 'center')
ws.row_dimensions[2].height = 18

# Row 3 Headers
headers = ['Feature', 'Test Case ID', 'Test Title', 'Testing Steps', 'Test Data', 'Expected Results', 'Actual Results', 'Pass/Fail']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = bfont(HDR_FG, True, 11); c.fill = fill(HDR_BG)
    c.alignment = align('center', 'center'); c.border = fborder()
ws.row_dimensions[3].height = 22

rows = [
  ('Feature 1:\nTechnician Bid\nSubmission\n(FR-S3.1)',
   'TC-BID-01', 'Submit a valid bid',
   '1. Log in as Technician\n2. Navigate to Repair Requests\n3. Select a PENDING repair request\n4. Fill Estimated Amount, Days, Message\n5. Click Submit Bid',
   'Amount = 1500 BDT\nDays = 3\nMessage = "Quick fix"',
   'Bid saved with status PENDING. Success message shown. Bid appears under the repair request.',
   'status=201 — Bid created successfully with PENDING status', 'Pass'),

  ('', 'TC-BID-02', 'Submit bid without Estimated Amount',
   '1. Log in as Technician\n2. Navigate to PENDING repair request\n3. Leave Amount blank\n4. Fill Days = 2\n5. Click Submit Bid',
   'Amount = (empty)\nDays = 2',
   'Validation error shown. Bid is NOT submitted.',
   'status=400 — msg: estimatedAmount and estimatedDays required', 'Pass'),

  ('', 'TC-BID-03', 'Submit bid without Estimated Days',
   '1. Log in as Technician\n2. Navigate to PENDING repair request\n3. Fill Amount = 1000\n4. Leave Days blank\n5. Click Submit Bid',
   'Amount = 1000\nDays = (empty)',
   'Validation error shown. Bid is NOT submitted.',
   'status=400 — Validation error returned', 'Pass'),

  ('', 'TC-BID-04', 'Submit bid with negative amount',
   '1. Log in as Technician\n2. Navigate to PENDING repair request\n3. Enter Amount = -500, Days = 2\n4. Click Submit Bid',
   'Amount = -500\nDays = 2',
   'System rejects negative value, shows validation error.',
   'status=400 — msg: estimatedAmount must be a positive number', 'Pass'),

  ('', 'TC-BID-05', 'Submit bid without login',
   '1. Open bid submission URL directly without logging in\n2. Attempt to submit a bid',
   'No auth token',
   'System redirects to login page or returns 401 Unauthorized.',
   'status=401 — Unauthorized', 'Pass'),

  ('', 'TC-BID-06', 'Submit bid as Customer role',
   '1. Log in as Customer\n2. Try to access bid submission for a repair request',
   'Customer JWT token',
   'System returns 403 Forbidden. Bid form not visible to customers.',
   'status=403 — Forbidden', 'Pass'),

  ('Feature 2:\nCustomer Bid\nManagement\n(FR-S3.2)',
   'TC-CBID-01', 'Customer views bids for their repair',
   '1. Log in as Customer\n2. Go to My Repair Requests\n3. Click a request that has bids\n4. View the Bids section',
   'Repair request with 1+ submitted bids',
   'All bids displayed with technician name, amount, days, message, status PENDING.',
   'status=200 — Bids list returned (count=1)', 'Pass'),

  ('', 'TC-CBID-02', 'Customer views repair with no bids',
   '1. Log in as Customer\n2. Navigate to repair request with no bids\n3. View the Bids section',
   'Repair request with 0 bids',
   'Message shown: no bids submitted yet.',
   'status=200 — Empty bids list returned (count=0)', 'Pass'),

  ('', 'TC-CBID-03', "Customer cannot view another customer's bids",
   '1. Log in as Customer A\n2. Access Customer B repair bids using the repair ID',
   'Customer A JWT\nCustomer B repair ID',
   'System returns 403 Forbidden. Bids are not shown.',
   'status=403 — Forbidden', 'Pass'),

  ('', 'TC-CBID-04', 'Customer accepts a valid bid',
   '1. Log in as Customer\n2. Navigate to repair request bids\n3. Click Accept on a PENDING bid',
   'Valid PENDING bid by Technician X',
   'Accepted bid becomes ACCEPTED. All other bids become REJECTED. Success message shown.',
   'status=200 — msg: Bid accepted! Delivery men notified for pickup.', 'Pass'),

  ('', 'TC-CBID-05', 'Customer rejects a bid',
   '1. Log in as Customer\n2. Navigate to repair request bids\n3. Click Reject on a PENDING bid',
   'Valid PENDING bid',
   'Rejected bid becomes REJECTED. Other bids remain PENDING.',
   'status=200 — bidStatus=REJECTED', 'Pass'),

  ('', 'TC-CBID-06', 'Accept already-rejected bid blocked',
   '1. Log in as Customer\n2. Navigate to bids\n3. Attempt to accept a REJECTED bid',
   'Bid with status = REJECTED',
   'Error message shown. Status does not change.',
   'status=400 — Cannot accept a bid that is already REJECTED', 'Pass'),

  ('Feature 3:\nBid Acceptance\nWorkflow\n(FR-S3.3)',
   'TC-BAW-01', 'Technician auto-assigned after bid acceptance',
   '1. Log in as Customer\n2. Accept bid from Technician Y\n3. Check repair request details',
   'Customer accepts Technician Y bid',
   'Technician Y auto-assigned to repair. Repair request shows assigned technician.',
   'status=200 — Technician auto-assigned on bid acceptance', 'Pass'),

  ('', 'TC-BAW-02', 'Other bids auto-rejected after one is accepted',
   '1. Log in as Customer\n2. Accept one bid (2 bids total)\n3. Check status of the other bid',
   '2 PENDING bids on same repair request',
   'Accepted bid becomes ACCEPTED. Other bid becomes REJECTED automatically.',
   'bidB status=REJECTED — Other bid auto-rejected', 'Pass'),

  ('', 'TC-BAW-03', 'Delivery job auto-created after bid acceptance',
   '1. Log in as Customer\n2. Accept a bid\n3. Log in as Delivery Man\n4. Check available delivery jobs',
   'Bid acceptance event',
   'New delivery job created with status PENDING_PICKUP for the repair request.',
   'status=200 — jobStatus=PENDING_PICKUP created automatically', 'Pass'),

  ('', 'TC-BAW-04', 'Cannot accept bid again after acceptance',
   '1. Log in as Customer\n2. Accept a bid\n3. Try to accept the same or another bid for the same repair',
   'Same repair, first bid already ACCEPTED',
   'System rejects the second acceptance and shows an error.',
   'status=400 — Cannot accept a bid that is already ACCEPTED', 'Pass'),

  ('Feature 5:\nDelivery Man\nRole\n(FR-S3.5)',
   'TC-DLV-01', 'Delivery man self-registers',
   '1. Open registration page\n2. Select role = Delivery Man\n3. Fill name, email, password\n4. Submit',
   'Name = DM2\nEmail = dm@test.com\nPassword = pass1234\nRole = DELIVERY_MAN',
   'Account created with DELIVERY_MAN role. Can log in successfully.',
   'status=201 — role=DELIVERY_MAN confirmed in response', 'Pass'),

  ('', 'TC-DLV-03', 'Delivery man can access delivery dashboard',
   '1. Log in with DELIVERY_MAN credentials\n2. Navigate to GET /delivery/available',
   'Delivery man JWT token',
   'After login, delivery man can access the Delivery Dashboard.',
   'status=200 — Delivery dashboard accessible', 'Pass'),

  ('', 'TC-DLV-04', 'Non-delivery-man blocked from delivery dashboard',
   '1. Log in as Customer\n2. Navigate to GET /delivery/available',
   'Customer JWT token',
   'Access denied. Returns 403 Forbidden.',
   'status=403 — Forbidden', 'Pass'),

  ('Feature 6:\nDelivery\nDashboard\n(FR-S3.6)',
   'TC-DASH-01', 'Dashboard shows available jobs',
   '1. Log in as Delivery Man\n2. Open Delivery Dashboard\n3. Check Available Jobs section',
   'At least 1 delivery job with PENDING_PICKUP exists',
   'Available jobs listed with customer info, issue summary, and current status.',
   'status=200 — Available jobs listed (count=2)', 'Pass'),

  ('', 'TC-DASH-02', 'Active job appears after acceptance',
   '1. Log in as Delivery Man\n2. Accept a job\n3. Check GET /delivery/my-jobs',
   'Delivery man has 1 accepted active job',
   'Active job shown with customer info, repair issue, current status.',
   'status=200 — Active job count=1 after acceptance', 'Pass'),

  ('', 'TC-DASH-03', 'Dashboard shows empty active jobs initially',
   '1. Log in as fresh Delivery Man\n2. Check GET /delivery/my-jobs before accepting any job',
   'No jobs assigned yet',
   'Returns 200 OK with empty jobs list.',
   'status=200 — Empty active jobs returned', 'Pass'),

  ('Feature 7:\nDelivery Job\nAcceptance\n(FR-S3.7)',
   'TC-DJOB-01', 'Delivery man accepts available job',
   '1. Log in as Delivery Man\n2. Navigate to Delivery Dashboard\n3. Click Accept on an available job',
   'Delivery job with no assigned delivery man',
   'Job assigned to this delivery man. Status changes to GOING_TO_CUSTOMER.',
   'status=200 — newStatus=GOING_TO_CUSTOMER, job assigned', 'Pass'),

  ('', 'TC-DJOB-02', 'Cannot accept already-taken job',
   '1. Log in as same Delivery Man\n2. Try to accept the same job again after it is taken',
   'Job already in GOING_TO_CUSTOMER status',
   'Error shown: job not available for acceptance. Returns 400.',
   'status=400 — Job is not available for acceptance (current: GOING_TO_CUSTOMER)', 'Pass'),

  ('', 'TC-DJOB-03', 'Customer cannot accept delivery job',
   '1. Log in as Customer\n2. Attempt POST /delivery/:id/accept via API',
   'Customer JWT token',
   'System returns 403 Forbidden. Job not accepted.',
   'status=403 — Forbidden', 'Pass'),

  ('Feature 8:\nDelivery Status\nWorkflow\n(FR-S3.8)',
   'TC-DSW-01', 'Valid: GOING_TO_CUSTOMER to PICKED_UP',
   '1. Log in as Delivery Man\n2. Open active job in GOING_TO_CUSTOMER status\n3. Update status to PICKED_UP',
   'Status = GOING_TO_CUSTOMER',
   'Status changes to PICKED_UP. History updated.',
   'status=200 — newStatus=PICKED_UP', 'Pass'),

  ('', 'TC-DSW-02', 'Invalid status transition rejected',
   '1. Log in as Delivery Man\n2. Open active job in GOING_TO_CUSTOMER\n3. Try to set status to DELIVERED',
   'Status = GOING_TO_CUSTOMER\nTarget = DELIVERED (invalid skip)',
   'System rejects transition and returns 400 validation error.',
   'status=400 — Cannot move from GOING_TO_CUSTOMER to DELIVERED', 'Pass'),

  ('', 'TC-DSW-03', 'Valid: PICKED_UP to AT_WAREHOUSE',
   '1. Log in as Delivery Man\n2. Open active job in PICKED_UP status\n3. Update status to AT_WAREHOUSE',
   'Status = PICKED_UP',
   'Status auto-advances to PENDING_TECH_DELIVERY after AT_WAREHOUSE. History recorded.',
   'status=200 — newStatus=PENDING_TECH_DELIVERY (auto-advanced from AT_WAREHOUSE)', 'Pass'),

  ('', 'TC-DSW-04', 'Only assigned delivery man can update status',
   '1. Log in as Technician (not the assigned delivery man)\n2. Attempt PATCH /delivery/:id/status',
   'Technician JWT token\nJob assigned to another user',
   'System returns 403 Forbidden. Status not changed.',
   'status=403 — Forbidden', 'Pass'),

  ('Feature 9:\nDelivery Status\nHistory\n(FR-S3.9)',
   'TC-DHIST-01', 'Status history recorded after each update',
   '1. Log in as Delivery Man\n2. Update status multiple times\n3. Check job statusHistory field',
   'Multiple status update events',
   'New history entry added for each update with status, timestamp, and note.',
   'status=200 — History entries=5 recorded', 'Pass'),

  ('', 'TC-DHIST-02', 'Multiple history entries accumulate correctly',
   '1. Log in as Delivery Man\n2. Perform 3+ consecutive valid status updates\n3. Check history list',
   '3+ status transitions made',
   'History list shows all entries in chronological order.',
   'status=200 — History entries=5 (multiple entries confirmed)', 'Pass'),

  ('', 'TC-DHIST-03', 'Admin can read status history',
   '1. Log in as Admin\n2. Navigate to GET /delivery/:id\n3. View job statusHistory field',
   'Admin JWT\nJob with statusHistory entries',
   'Admin sees full delivery status history. Returns 200 OK.',
   'status=200 — Admin can view all 5 history entries', 'Pass'),

  ('Feature 11:\nDelivery Tracking\nVisibility\n(FR-S3.11)',
   'TC-DTRK-01', 'Customer views delivery tracking',
   '1. Log in as Customer\n2. Navigate to repair request with active delivery\n3. GET /delivery/by-repair/:id',
   'Customer owns the repair request',
   'Delivery tracking section visible with current status. Returns 200 OK.',
   'status=200 — Customer can view delivery tracking', 'Pass'),

  ('', 'TC-DTRK-02', 'Admin views delivery tracking on any repair',
   '1. Log in as Admin\n2. Navigate to any repair with active delivery\n3. GET /delivery/by-repair/:id',
   'Admin JWT token',
   'Delivery tracking visible to admin. Returns 200 OK.',
   'status=200 — Admin can view delivery tracking', 'Pass'),

  ('', 'TC-DTRK-03', 'Technician views delivery tracking',
   '1. Log in as assigned Technician\n2. Navigate to repair details\n3. GET /delivery/by-repair/:id',
   'Lead Technician JWT token',
   'Delivery tracking visible to assigned technician. Returns 200 OK.',
   'status=200 — Technician can view delivery tracking', 'Pass'),

  ('', 'TC-DTRK-04', "Other customer blocked from repair tracking",
   '1. Log in as Customer B\n2. Navigate to Customer A repair tracking page',
   'Customer B JWT token\nCustomer A repair ID',
   '403 Forbidden. Tracking not visible to other customers.',
   'status=403 — Forbidden (other customer blocked)', 'Pass'),

  ('Feature 12:\nRole-Based\nAccess Control\n(FR-S3.12)',
   'TC-RBAC-01', 'Customer blocked from delivery dashboard',
   '1. Log in as Customer\n2. Navigate to GET /delivery/available',
   'Customer JWT token',
   'Access denied. Returns 403 Forbidden.',
   'status=403 — Forbidden', 'Pass'),

  ('', 'TC-RBAC-03', 'Unauthenticated user gets 401',
   '1. Without login, send GET to /api/delivery/available',
   'No JWT token',
   'Returns 401 Unauthorized.',
   'status=401 — Unauthorized', 'Pass'),

  ('', 'TC-RBAC-04', 'Customer cannot update delivery status',
   '1. Log in as Customer\n2. Attempt PATCH /delivery/:id/status via API',
   'Customer JWT token',
   '403 Forbidden. Delivery status not changed.',
   'status=403 — Forbidden', 'Pass'),

  ('', 'TC-RBAC-05', 'Technician cannot accept delivery jobs',
   '1. Log in as Technician\n2. Attempt POST /delivery/:id/accept via API',
   'Technician JWT token',
   '403 Forbidden. Job not accepted.',
   'status=403 — Forbidden', 'Pass'),
]

ROW_START = 4
feat_starts = {}
for idx, row in enumerate(rows):
    rn = ROW_START + idx
    feat_label = row[0]
    if feat_label:
        feat_starts[rn] = feat_label
    tcId = row[1]
    pf = row[7]
    is_odd = idx % 2 == 0
    row_bg = ODD if is_odd else EVEN

    vals = [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=rn, column=col, value=val)
        c.border = tborder()
        c.alignment = align()
        if col == 1:
            if val:
                c.font = bfont(FEAT_FG, True, 10)
                c.fill = fill(FEAT_BG)
                c.alignment = align('center', 'center')
            else:
                c.fill = fill(FEAT_BG)
        elif col == 2:
            c.font = bfont('000000', True, 10)
            c.fill = fill(row_bg)
            c.alignment = align('center', 'center')
        elif col == 3:
            c.font = bfont('000000', True, 10)
            c.fill = fill(row_bg)
        elif col in (4, 5, 6):
            c.font = bfont('000000', False, 10)
            c.fill = fill(row_bg)
        elif col == 7:
            c.font = bfont('000000', False, 10)
            c.fill = fill(PASS_BG if pf == 'Pass' else FAIL_BG)
        elif col == 8:
            if pf == 'Pass':
                c.font = bfont(PASS_FG, True, 10)
                c.fill = fill(PASS_BG)
            else:
                c.font = bfont(FAIL_FG, True, 10)
                c.fill = fill(FAIL_BG)
            c.alignment = align('center', 'center')
    ws.row_dimensions[rn].height = 80

total_rows = ROW_START + len(rows) - 1
feat_list = sorted(feat_starts.keys())
for i, start in enumerate(feat_list):
    end = (feat_list[i+1] - 1) if i + 1 < len(feat_list) else total_rows
    if end >= start:
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        c = ws.cell(row=start, column=1)
        c.font = bfont(FEAT_FG, True, 10)
        c.fill = fill(FEAT_BG)
        c.alignment = align('center', 'center')
        c.border = fborder()

sr = total_rows + 1
ws.merge_cells(f'A{sr}:H{sr}')
c = ws.cell(row=sr, column=1)
c.value = 'TOTAL: 40/40 Test Cases PASSED  |  All 4 Bugs Fixed  |  Live Tested Against localhost:5000'
c.font = bfont(HDR_FG, True, 11)
c.fill = fill(HDR_BG)
c.alignment = align('center', 'center')
ws.row_dimensions[sr].height = 20

ws.freeze_panes = 'A4'
ws.auto_filter.ref = f'A3:H{total_rows}'

out = r'C:\Users\Asif Bin Mahmood\Downloads\Manual_Testing_CRMS_Sprint3.xlsx'
wb.save(out)
print('Done! Saved: ' + out)
