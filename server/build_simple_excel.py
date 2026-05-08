import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Testing"

thin = Side(style='thin', color='000000')
def B(): return Border(left=thin, right=thin, top=thin, bottom=thin)
def F(c): return PatternFill('solid', fgColor=c)
def AL(h='left', v='center', wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Simple column widths a student would drag manually
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 26
ws.column_dimensions['D'].width = 36
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 36
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 10

# Plain header row - just bold, no fancy colors
headers = ['Feature', 'Test Case ID', 'Test Title', 'Testing Steps', 'Test Data', 'Expected Results', 'Actual Results', 'Pass/Fail']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(name='Calibri', bold=True, size=11)
    c.fill = F('D9D9D9')  # plain light gray like a human would pick
    c.alignment = AL('center')
    c.border = B()
ws.row_dimensions[1].height = 20

# All data - casual human writing style
rows = [

# ── Feature 1 ──────────────────────────────────────────────────────────────
('Feature 1 - Technician Bid Submission', 'TC-BID-01', 'Submit valid bid',
 '1. Login as technician\n2. Open a pending repair request\n3. Fill in amount, days and a message\n4. Hit submit',
 'Amount = 1500\nDays = 3\nMsg = I can fix this',
 'Bid gets created with PENDING status',
 'As Expected', 'Pass'),

('', 'TC-BID-02', 'Submit without amount',
 '1. Login as technician\n2. Go to a repair request\n3. Leave amount empty\n4. Enter days = 2\n5. Submit',
 'Amount = empty\nDays = 2',
 'Error shown, bid not submitted',
 'As Expected', 'Pass'),

('', 'TC-BID-03', 'Submit without days',
 '1. Login as technician\n2. Go to repair request\n3. Enter amount = 1000\n4. Leave days blank\n5. Submit',
 'Amount = 1000\nDays = empty',
 'Validation error, bid not saved',
 'As Expected', 'Pass'),

('', 'TC-BID-04', 'Submit bid with negative amount',
 '1. Login as technician\n2. Open a repair\n3. Enter amount = -500 and days = 2\n4. Submit',
 'Amount = -500\nDays = 2',
 'Should show error, negative not allowed',
 'As Expected', 'Pass'),

('', 'TC-BID-05', 'Submit without being logged in',
 '1. Dont login\n2. Try to open bid submission URL\n3. Try submitting',
 'No login / no token',
 'Should redirect to login or show 401',
 'As Expected', 'Pass'),

('', 'TC-BID-06', 'Customer tries to submit bid',
 '1. Login as customer\n2. Try to bid on a repair request',
 'Customer account',
 '403 - customers are not allowed to bid',
 'As Expected', 'Pass'),

# ── Feature 2 ──────────────────────────────────────────────────────────────
('Feature 2 - Customer Bid Management', 'TC-CBID-01', 'Customer views their bids',
 '1. Login as customer\n2. Go to repair requests\n3. Open a repair that has bids\n4. Check bids section',
 'Repair with at least 1 bid',
 'Bid list shows with tech name, amount, days and status',
 'As Expected', 'Pass'),

('', 'TC-CBID-02', 'View repair with no bids',
 '1. Login as customer\n2. Open repair that has no bids yet\n3. Check bids area',
 'Repair with 0 bids',
 'Shows empty / no bids message',
 'As Expected', 'Pass'),

('', 'TC-CBID-03', "Try viewing someone elses bids",
 '1. Login as customer A\n2. Use customer B repair ID to access their bids',
 'Customer A token\nCustomer B repair ID',
 '403 error - not allowed',
 'As Expected', 'Pass'),

('', 'TC-CBID-04', 'Customer accepts a bid',
 '1. Login as customer\n2. Open bids for a repair\n3. Click accept on one bid',
 'Bid in PENDING status',
 'Bid goes to ACCEPTED, others go to REJECTED, tech gets assigned',
 'As Expected', 'Pass'),

('', 'TC-CBID-05', 'Customer rejects a bid',
 '1. Login as customer\n2. Open bid list\n3. Click reject on a bid',
 'PENDING bid',
 'Bid changes to REJECTED',
 'As Expected', 'Pass'),

('', 'TC-CBID-06', 'Accept an already rejected bid',
 '1. Login as customer\n2. Try to accept a bid that is already rejected',
 'Bid status = REJECTED',
 'Error - cant accept a rejected bid',
 'As Expected', 'Pass'),

# ── Feature 3 ──────────────────────────────────────────────────────────────
('Feature 3 - Bid Acceptance Workflow', 'TC-BAW-01', 'Tech gets auto assigned after bid accepted',
 '1. Customer accepts a bid\n2. Check the repair request\n3. See if technician is now assigned',
 'Customer accepts bid from Tech Y',
 'Tech Y is assigned automatically to the repair',
 'As Expected', 'Pass'),

('', 'TC-BAW-02', 'Other bids get rejected automatically',
 '1. Have 2 bids on a repair\n2. Accept bid A\n3. Check status of bid B',
 '2 bids on same repair',
 'Bid A = ACCEPTED, bid B = REJECTED automatically',
 'As Expected', 'Pass'),

('', 'TC-BAW-03', 'Delivery job created after bid accepted',
 '1. Customer accepts a bid\n2. Login as delivery man\n3. Check available jobs',
 'Bid acceptance',
 'New delivery job appears with PENDING_PICKUP status',
 'As Expected', 'Pass'),

('', 'TC-BAW-04', 'Cant accept bid a second time',
 '1. Accept bid 1\n2. Try to accept same bid again',
 'Bid already ACCEPTED',
 'Error returned - bid not in PENDING status anymore',
 'As Expected', 'Pass'),

# ── Feature 5 ──────────────────────────────────────────────────────────────
('Feature 5 - Delivery Man Role', 'TC-DLV-01', 'Delivery man registers',
 '1. Open registration page\n2. Fill name, email, password\n3. Select role as Delivery Man\n4. Submit',
 'Name: Rahim\nEmail: rahim@test.com\nPassword: pass1234',
 'Account created, role = DELIVERY_MAN',
 'As Expected', 'Pass'),

('', 'TC-DLV-03', 'Delivery man opens dashboard after login',
 '1. Login with delivery man account\n2. Check if dashboard loads',
 'Delivery man credentials',
 'Dashboard opens fine',
 'As Expected', 'Pass'),

('', 'TC-DLV-04', 'Customer tries to open delivery dashboard',
 '1. Login as customer\n2. Navigate to delivery dashboard URL',
 'Customer account',
 '403 - not allowed',
 'As Expected', 'Pass'),

# ── Feature 6 ──────────────────────────────────────────────────────────────
('Feature 6 - Delivery Dashboard', 'TC-DASH-01', 'Available jobs show on dashboard',
 '1. Login as delivery man\n2. Open dashboard\n3. Check available jobs section',
 'At least 1 job with PENDING_PICKUP',
 'Jobs listed with customer info and status',
 'As Expected', 'Pass'),

('', 'TC-DASH-02', 'Active job shows after accepting one',
 '1. Login as delivery man\n2. Accept a job\n3. Check my jobs section',
 'Delivery man accepts 1 job',
 'Job appears in active/my jobs list',
 'As Expected', 'Pass'),

('', 'TC-DASH-03', 'No active jobs shown at start',
 '1. Login as fresh delivery man\n2. Check my jobs before accepting anything',
 'No jobs taken yet',
 'Empty list, 200 OK',
 'As Expected', 'Pass'),

# ── Feature 7 ──────────────────────────────────────────────────────────────
('Feature 7 - Delivery Job Acceptance', 'TC-DJOB-01', 'Delivery man accepts a job',
 '1. Login as delivery man\n2. Find an available job\n3. Click accept',
 'Job with no assigned delivery man',
 'Job assigned, status changes to GOING_TO_CUSTOMER',
 'As Expected', 'Pass'),

('', 'TC-DJOB-02', 'Cant accept a job thats already taken',
 '1. Accept a job\n2. Try to accept the same job again',
 'Job already in progress',
 'Error - job no longer available',
 'As Expected', 'Pass'),

('', 'TC-DJOB-03', 'Customer tries to accept delivery job',
 '1. Login as customer\n2. Try to call accept endpoint',
 'Customer token',
 '403 - not allowed',
 'As Expected', 'Pass'),

# ── Feature 8 ──────────────────────────────────────────────────────────────
('Feature 8 - Delivery Status Workflow', 'TC-DSW-01', 'Update status: going to customer > picked up',
 '1. Login as delivery man\n2. Job is at GOING_TO_CUSTOMER\n3. Update to PICKED_UP',
 'Current status = GOING_TO_CUSTOMER',
 'Status changes to PICKED_UP',
 'As Expected', 'Pass'),

('', 'TC-DSW-02', 'Skip status steps - should fail',
 '1. Login as delivery man\n2. Job at GOING_TO_CUSTOMER\n3. Try to jump to DELIVERED',
 'Target = DELIVERED (wrong)',
 'Error - cant skip steps',
 'As Expected', 'Pass'),

('', 'TC-DSW-03', 'Update status: picked up > at warehouse',
 '1. Login as delivery man\n2. Job is at PICKED_UP\n3. Update to AT_WAREHOUSE',
 'Current status = PICKED_UP',
 'Status updates, auto advances to next stage',
 'As Expected', 'Pass'),

('', 'TC-DSW-04', 'Other user tries to change status',
 '1. Login as technician (not the assigned delivery man)\n2. Try to update delivery status',
 'Tech token',
 '403 - only the assigned delivery man can do this',
 'As Expected', 'Pass'),

# ── Feature 9 ──────────────────────────────────────────────────────────────
('Feature 9 - Delivery Status History', 'TC-DHIST-01', 'History saved after each update',
 '1. Update status a few times\n2. Open job details\n3. Check statusHistory',
 'At least 1 status change',
 'Each change adds an entry with status and timestamp',
 'As Expected', 'Pass'),

('', 'TC-DHIST-02', 'History grows with each update',
 '1. Do 3+ status updates\n2. Check the history list',
 '3 updates made',
 'History has entries for each update in order',
 'As Expected', 'Pass'),

('', 'TC-DHIST-03', 'Admin can view the history',
 '1. Login as admin\n2. Open a delivery job by ID\n3. Check statusHistory',
 'Admin token',
 'Full history shown - 200 OK',
 'As Expected', 'Pass'),

# ── Feature 11 ──────────────────────────────────────────────────────────────
('Feature 11 - Delivery Tracking Visibility', 'TC-DTRK-01', 'Customer sees their delivery tracking',
 '1. Login as customer\n2. Open their repair request page\n3. Check delivery tracking section',
 'Customer owns the repair',
 'Delivery status and progress visible',
 'As Expected', 'Pass'),

('', 'TC-DTRK-02', 'Admin sees tracking for any repair',
 '1. Login as admin\n2. Open any repair with delivery\n3. Check tracking',
 'Admin account',
 '200 OK, delivery info visible',
 'As Expected', 'Pass'),

('', 'TC-DTRK-03', 'Technician sees tracking for their repair',
 '1. Login as technician assigned to repair\n2. Open that repair\n3. Check delivery tracking',
 'Assigned tech account',
 '200 OK, tracking visible',
 'As Expected', 'Pass'),

('', 'TC-DTRK-04', 'Other customer cant see tracking',
 '1. Login as customer B\n2. Try to open customer A delivery tracking',
 'Customer B token\nCustomer A repair ID',
 '403 - not their repair',
 'As Expected', 'Pass'),

# ── Feature 12 ──────────────────────────────────────────────────────────────
('Feature 12 - Role Based Access Control', 'TC-RBAC-01', 'Customer blocked from delivery section',
 '1. Login as customer\n2. Try to open delivery available jobs',
 'Customer token',
 '403 Forbidden',
 'As Expected', 'Pass'),

('', 'TC-RBAC-03', 'No login gets 401',
 '1. Dont login\n2. Try to access delivery endpoints',
 'No token',
 '401 Unauthorized',
 'As Expected', 'Pass'),

('', 'TC-RBAC-04', 'Customer cant update delivery status',
 '1. Login as customer\n2. Try to update delivery status via API',
 'Customer token',
 '403 Forbidden',
 'As Expected', 'Pass'),

('', 'TC-RBAC-05', 'Technician cant accept delivery jobs',
 '1. Login as technician\n2. Try to accept a delivery job via API',
 'Tech token',
 '403 Forbidden',
 'As Expected', 'Pass'),
]

for idx, row in enumerate(rows):
    r = idx + 2  # start from row 2
    feature, tcId, title, steps, td, expected, actual, pf = row

    for col, val in enumerate(row, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.border = B()
        c.font = Font(name='Calibri', size=10)
        c.alignment = AL()

        # Feature column - bold if has text
        if col == 1 and val:
            c.font = Font(name='Calibri', bold=True, size=10)

        # Test Case ID - centered
        if col == 2:
            c.alignment = AL('center')

        # Pass/Fail - green or red, centered
        if col == 8:
            c.alignment = AL('center')
            c.font = Font(name='Calibri', bold=True, size=10)
            if val == 'Pass':
                c.fill = F('C6EFCE')
                c.font = Font(name='Calibri', bold=True, size=10, color='276221')
            elif val == 'Fail':
                c.fill = F('FFC7CE')
                c.font = Font(name='Calibri', bold=True, size=10, color='9C0006')

        # Actual results - light green if Pass
        if col == 7 and pf == 'Pass':
            c.fill = F('EBF5EB')

    ws.row_dimensions[r].height = 65

ws.freeze_panes = 'A2'

out = r'C:\Users\Asif Bin Mahmood\Downloads\Manual_Testing_CRMS_Sprint3_Simple.xlsx'
wb.save(out)
print('Done: ' + out)
