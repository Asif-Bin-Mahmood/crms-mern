import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Testing"

thin = Side(style='thin', color='000000')
def B(): return Border(left=thin, right=thin, top=thin, bottom=thin)
def F(c): return PatternFill('solid', fgColor=c)
def AL(h='left', v='center', wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 26
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 26
ws.column_dimensions['F'].width = 38
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 10

headers = ['Feature', 'Test Case ID', 'Test Title', 'Testing Steps', 'Test Data', 'Expected Results', 'Actual Results', 'Pass/Fail']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(name='Calibri', bold=True, size=11)
    c.fill = F('D9D9D9')
    c.alignment = AL('center')
    c.border = B()
ws.row_dimensions[1].height = 20

rows = [

# ── Feature 1 ──────────────────────────────────────────────────────────────
('Feature 1 - Technician Bid Submission', 'TC-BID-01', 'Submit a valid bid',
 '1. Open browser, go to http://localhost:5173\n2. Login with technician account\n3. Click on "Repair Requests" from the menu\n4. Open a PENDING repair request\n5. Fill in estimated amount, days and a message\n6. Click "Submit Bid" button',
 'Amount: 1500\nDays: 3\nMessage: I can fix this',
 'Bid appears in the list with PENDING status. Success message shown on screen.',
 'As Expected', 'Pass'),

('', 'TC-BID-02', 'Submit bid with empty amount',
 '1. Login as technician\n2. Open a PENDING repair request\n3. Leave the amount field empty\n4. Enter days = 2\n5. Click Submit Bid',
 'Amount: (empty)\nDays: 2',
 'Error message shown below the amount field. Bid is not submitted.',
 'As Expected', 'Pass'),

('', 'TC-BID-03', 'Submit bid with empty days',
 '1. Login as technician\n2. Open a PENDING repair request\n3. Enter amount = 1000\n4. Leave days field empty\n5. Click Submit Bid',
 'Amount: 1000\nDays: (empty)',
 'Error message shown. Bid is not saved.',
 'As Expected', 'Pass'),

('', 'TC-BID-04', 'Submit bid with negative amount',
 '1. Login as technician\n2. Open repair request\n3. Type -500 in the amount field\n4. Enter days = 2\n5. Click Submit Bid',
 'Amount: -500\nDays: 2',
 'Validation error shown. Form does not submit with negative value.',
 'As Expected', 'Pass'),

('', 'TC-BID-05', 'Try to bid without logging in',
 '1. Open browser without logging in\n2. Try to navigate to a repair request\n3. Try to submit a bid',
 'Not logged in',
 'User is redirected to login page. Bid form is not accessible.',
 'As Expected', 'Pass'),

('', 'TC-BID-06', 'Customer account tries to bid',
 '1. Login with a customer account\n2. Open a repair request\n3. Check if bid form is available',
 'Customer email and password',
 'Bid form is not visible. Customer cannot submit bids.',
 'As Expected', 'Pass'),

# ── Feature 2 ──────────────────────────────────────────────────────────────
('Feature 2 - Customer Bid Management', 'TC-CBID-01', 'Customer views bids on their repair',
 '1. Open browser, go to http://localhost:5173\n2. Login as customer\n3. Go to "My Repairs" from menu\n4. Click on a repair request\n5. Scroll to bids section',
 'Repair has at least 1 bid submitted by a technician',
 'Bid list is visible showing technician name, amount, days and status',
 'As Expected', 'Pass'),

('', 'TC-CBID-02', 'View repair with no bids yet',
 '1. Login as customer\n2. Go to My Repairs\n3. Click on a repair that has no bids\n4. Open bids section',
 'Repair with 0 bids',
 'Empty message shown like "No bids yet"',
 'As Expected', 'Pass'),

('', 'TC-CBID-03', "Try to view another customer's bids",
 '1. Login as customer A\n2. Manually type another customer repair URL in the browser\n3. Try to open their bids',
 'Customer A account\nCustomer B repair URL',
 'Access denied or redirected. Other customers bids are not visible.',
 'As Expected', 'Pass'),

('', 'TC-CBID-04', 'Customer accepts a bid',
 '1. Login as customer\n2. Go to My Repairs\n3. Open a repair with bids\n4. Click "Accept" on one of the bids',
 'PENDING bid by a technician',
 'Accepted bid shows ACCEPTED status. Others show REJECTED. Technician gets assigned.',
 'As Expected', 'Pass'),

('', 'TC-CBID-05', 'Customer rejects a bid',
 '1. Login as customer\n2. Go to My Repairs and open a repair\n3. Click "Reject" on a PENDING bid',
 'PENDING bid',
 'Bid status changes to REJECTED on the screen',
 'As Expected', 'Pass'),

('', 'TC-CBID-06', 'Try to accept a rejected bid',
 '1. Login as customer\n2. Open bids section on a repair\n3. Try to click accept on a bid that is already REJECTED',
 'Bid already in REJECTED status',
 'Error message shown. Status stays REJECTED.',
 'As Expected', 'Pass'),

# ── Feature 3 ──────────────────────────────────────────────────────────────
('Feature 3 - Bid Acceptance Workflow', 'TC-BAW-01', 'Technician assigned after customer accepts bid',
 '1. Login as customer\n2. Accept a bid from technician Y\n3. Logout and login as admin\n4. Check the repair request to see assigned technician',
 'Customer accepts bid from Tech Y',
 'Tech Y is shown as the assigned technician on the repair request',
 'As Expected', 'Pass'),

('', 'TC-BAW-02', 'Other bids become rejected automatically',
 '1. Login as customer\n2. Open a repair with 2 bids\n3. Accept one bid\n4. Check status of the other bid',
 '2 bids on same repair',
 'Accepted bid = ACCEPTED. The other bid automatically shows REJECTED.',
 'As Expected', 'Pass'),

('', 'TC-BAW-03', 'Delivery job appears after bid accepted',
 '1. Login as customer and accept a bid\n2. Logout\n3. Login as delivery man\n4. Open delivery dashboard\n5. Check available jobs',
 'Bid acceptance by customer',
 'New delivery job appears in the available jobs list with PENDING PICKUP status',
 'As Expected', 'Pass'),

('', 'TC-BAW-04', 'Cannot accept another bid after one is accepted',
 '1. Login as customer\n2. Accept one bid on a repair\n3. Try to accept another bid on the same repair',
 'Already accepted 1 bid',
 'Error message shown. Second accept is blocked.',
 'As Expected', 'Pass'),

# ── Feature 5 ──────────────────────────────────────────────────────────────
('Feature 5 - Delivery Man Role', 'TC-DLV-01', 'Delivery man creates an account',
 '1. Open browser and go to registration page\n2. Fill in name, email, password\n3. Select role = Delivery Man\n4. Click Register',
 'Name: Rahim\nEmail: rahim@test.com\nPassword: pass1234\nRole: Delivery Man',
 'Account is created. User can login with these details.',
 'As Expected', 'Pass'),

('', 'TC-DLV-03', 'Delivery man redirected to delivery dashboard',
 '1. Open browser\n2. Login with delivery man email and password\n3. Watch where the page goes after login',
 'Delivery man credentials',
 'After login, user is taken to the delivery dashboard automatically',
 'As Expected', 'Pass'),

('', 'TC-DLV-04', 'Customer cannot access delivery dashboard',
 '1. Login as customer\n2. Try typing the delivery dashboard URL in the browser',
 'Customer account',
 'Page shows access denied or redirects to customer home. Dashboard not accessible.',
 'As Expected', 'Pass'),

# ── Feature 6 ──────────────────────────────────────────────────────────────
('Feature 6 - Delivery Dashboard', 'TC-DASH-01', 'Available jobs are shown on dashboard',
 '1. Login as delivery man\n2. Look at the delivery dashboard\n3. Check the available jobs section',
 'At least 1 pending delivery job in the system',
 'Available jobs list shows with customer info, repair issue and current status',
 'As Expected', 'Pass'),

('', 'TC-DASH-02', 'Active job shows after delivery man accepts it',
 '1. Login as delivery man\n2. Accept an available job\n3. Check the active/my jobs section on dashboard',
 'Delivery man accepts 1 job',
 'That job moves to the active jobs section on the dashboard',
 'As Expected', 'Pass'),

('', 'TC-DASH-03', 'Dashboard shows empty when no jobs assigned',
 '1. Login as a delivery man who has not accepted any job\n2. Open the dashboard\n3. Check active jobs section',
 'Fresh delivery man account',
 'Active jobs section shows empty or "no active jobs" message',
 'As Expected', 'Pass'),

# ── Feature 7 ──────────────────────────────────────────────────────────────
('Feature 7 - Delivery Job Acceptance', 'TC-DJOB-01', 'Delivery man accepts a job',
 '1. Login as delivery man\n2. Open delivery dashboard\n3. Find an available job in the list\n4. Click "Accept Job" button',
 'Available job with no assigned delivery man',
 'Job is now assigned to this delivery man. Status changes to GOING TO CUSTOMER. Job moves to active jobs.',
 'As Expected', 'Pass'),

('', 'TC-DJOB-02', 'Same job cannot be accepted by another delivery man',
 '1. Delivery man A accepts a job\n2. Login as delivery man B\n3. Check if the same job is still in available jobs\n4. Try to accept it',
 'Job already accepted by DM A',
 'Job is no longer in available list. Cannot be accepted again.',
 'As Expected', 'Pass'),

('', 'TC-DJOB-03', 'Customer cannot see accept job button',
 '1. Login as customer\n2. Try to navigate to delivery dashboard or any delivery page',
 'Customer account',
 'Delivery job accept option is not visible. Access is blocked.',
 'As Expected', 'Pass'),

# ── Feature 8 ──────────────────────────────────────────────────────────────
('Feature 8 - Delivery Status Workflow', 'TC-DSW-01', 'Update status step by step',
 '1. Login as delivery man\n2. Open an active job\n3. Current status is GOING TO CUSTOMER\n4. Click update status button\n5. Select PICKED UP\n6. Confirm',
 'Job status = GOING TO CUSTOMER',
 'Status changes to PICKED UP on the screen. Progress tracker updates.',
 'As Expected', 'Pass'),

('', 'TC-DSW-02', 'Cannot skip to a later status',
 '1. Login as delivery man\n2. Open active job (status = GOING TO CUSTOMER)\n3. Try to select DELIVERED directly in status update',
 'Trying to skip intermediate steps',
 'Error message shown. Invalid status change is rejected.',
 'As Expected', 'Pass'),

('', 'TC-DSW-03', 'Update: picked up to at warehouse',
 '1. Login as delivery man\n2. Open job at PICKED UP status\n3. Update status to AT WAREHOUSE\n4. Confirm',
 'Job status = PICKED UP',
 'Status updates. Progress tracker moves forward on screen.',
 'As Expected', 'Pass'),

('', 'TC-DSW-04', 'Other delivery man cannot change status',
 '1. Login as a different delivery man (not the assigned one)\n2. Try to update status on a job assigned to someone else',
 'Different delivery man account',
 'Error shown. Status update is blocked. Only assigned DM can update.',
 'As Expected', 'Pass'),

# ── Feature 9 ──────────────────────────────────────────────────────────────
('Feature 9 - Delivery Status History', 'TC-DHIST-01', 'Status history shows after each update',
 '1. Login as delivery man\n2. Update status on a job\n3. Open job details\n4. Look for status history or timeline section',
 'At least 1 status update done',
 'History entry visible with the status, time and note',
 'As Expected', 'Pass'),

('', 'TC-DHIST-02', 'History grows with every status update',
 '1. Do 3 or more status updates on the same job\n2. Open job detail page\n3. Check the history list',
 '3 status updates made',
 'Each update appears as a new entry in the history. Entries shown in time order.',
 'As Expected', 'Pass'),

('', 'TC-DHIST-03', 'Admin sees full status history',
 '1. Login as admin\n2. Open a repair request detail page\n3. Check the delivery tracking or history section',
 'Admin account',
 'Full delivery history is visible to admin',
 'As Expected', 'Pass'),

# ── Feature 11 ──────────────────────────────────────────────────────────────
('Feature 11 - Delivery Tracking Visibility', 'TC-DTRK-01', 'Customer sees delivery progress on their repair page',
 '1. Login as customer\n2. Go to My Repairs\n3. Click on a repair that has an active delivery\n4. Scroll down to delivery tracking section',
 'Customer repair with active delivery job',
 'Delivery status and progress tracker is visible on the repair detail page',
 'As Expected', 'Pass'),

('', 'TC-DTRK-02', 'Admin sees delivery tracking on any repair',
 '1. Login as admin\n2. Open any repair request in the system\n3. Check delivery tracking section',
 'Admin account',
 'Delivery tracking info is visible to admin',
 'As Expected', 'Pass'),

('', 'TC-DTRK-03', 'Technician sees delivery tracking for assigned repair',
 '1. Login as technician\n2. Open a repair assigned to them\n3. Check delivery tracking section on that page',
 'Technician assigned to this repair',
 'Delivery tracking section is visible',
 'As Expected', 'Pass'),

('', 'TC-DTRK-04', 'Customer cannot see another customers repair tracking',
 '1. Login as customer B\n2. Manually type the URL of customer A repair page in the browser\n3. Check if delivery tracking is visible',
 'Customer B account\nCustomer A repair page URL',
 'Page shows access denied or redirects. Tracking is not visible.',
 'As Expected', 'Pass'),

# ── Feature 12 ──────────────────────────────────────────────────────────────
('Feature 12 - Role Based Access Control', 'TC-RBAC-01', 'Customer cannot open delivery dashboard',
 '1. Login as customer\n2. Type the delivery dashboard URL in the browser address bar\n3. Press enter',
 'Customer account',
 'Page shows access denied or redirects to customer home page',
 'As Expected', 'Pass'),

('', 'TC-RBAC-03', 'Cannot access app without logging in',
 '1. Open a new private/incognito browser window\n2. Try to go directly to a protected page like the delivery dashboard',
 'Not logged in at all',
 'User is redirected to the login page',
 'As Expected', 'Pass'),

('', 'TC-RBAC-04', 'Customer cannot change delivery status',
 '1. Login as customer\n2. Try to navigate to any delivery status update page or button',
 'Customer account',
 'Status update option is not visible or accessible to customers',
 'As Expected', 'Pass'),

('', 'TC-RBAC-05', 'Technician cannot accept delivery jobs',
 '1. Login as technician\n2. Try to navigate to delivery dashboard or find an accept job button',
 'Technician account',
 'Accept job button is not visible. Delivery dashboard is not accessible to technicians.',
 'As Expected', 'Pass'),
]

for idx, row in enumerate(rows):
    r = idx + 2
    feature, tcId, title, steps, td, expected, actual, pf = row
    for col, val in enumerate(row, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.border = B()
        c.font = Font(name='Calibri', size=10)
        c.alignment = AL()
        if col == 1 and val:
            c.font = Font(name='Calibri', bold=True, size=10)
        if col == 2:
            c.alignment = AL('center')
        if col == 8:
            c.alignment = AL('center')
            if val == 'Pass':
                c.fill = F('C6EFCE')
                c.font = Font(name='Calibri', bold=True, size=10, color='276221')
            elif val == 'Fail':
                c.fill = F('FFC7CE')
                c.font = Font(name='Calibri', bold=True, size=10, color='9C0006')
        if col == 7 and pf == 'Pass':
            c.fill = F('EBF5EB')
    ws.row_dimensions[r].height = 72

ws.freeze_panes = 'A2'
out = r'C:\Users\Asif Bin Mahmood\Downloads\Manual_Testing_CRMS_Sprint3_Final.xlsx'
wb.save(out)
print('Done: ' + out)
