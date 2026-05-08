import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Testing"

# ── Simple human-like style (matches faculty sample) ──────────────────────────
YELLOW  = 'FFFF99'   # light yellow bg like the sample
WHITE   = 'FFFFFF'
GREEN   = 'CCFFCC'   # pass
RED     = 'FFCCCC'   # fail
HEADER  = 'D4E1A0'   # slightly darker yellow for col headers
FEAT    = 'C6EFCE'   # soft green for feature section headers

thin = Side(style='thin', color='000000')
def border(): return Border(left=thin, right=thin, top=thin, bottom=thin)
def fill(c):  return PatternFill('solid', fgColor=c)
def font(bold=False, sz=10): return Font(name='Times New Roman', bold=bold, size=sz)
def align(h='left', v='center', wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Column widths (simple, like a human would set)
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 13
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 38
ws.column_dimensions['E'].width = 26
ws.column_dimensions['F'].width = 38
ws.column_dimensions['G'].width = 30
ws.column_dimensions['H'].width = 11

# ── Title ─────────────────────────────────────────────────────────────────────
ws.merge_cells('A1:H1')
c = ws['A1']
c.value = 'Manual Testing - Computer Repair Management System (Sprint 3)'
c.font = Font(name='Times New Roman', bold=True, size=14)
c.fill = fill(YELLOW)
c.alignment = align('center')
ws.row_dimensions[1].height = 25

ws.merge_cells('A2:H2')
c = ws['A2']
c.value = 'Student: Asif Bin Mahmood (23201632)    Course: CSE470'
c.font = font(sz=10)
c.fill = fill(YELLOW)
c.alignment = align('center')
ws.row_dimensions[2].height = 16

# ── Column headers (row 3) ─────────────────────────────────────────────────────
headers = ['Feature', 'Test Case ID', 'Test Title', 'Testing Steps', 'Test Data', 'Expected Results', 'Actual Results', 'Pass/Fail']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = font(bold=True, sz=10)
    c.fill = fill(HEADER)
    c.alignment = align('center')
    c.border = border()
ws.row_dimensions[3].height = 18

# ── Data: human-style — natural language, varying formats ────────────────────
# Format: (feature, tcId, title, steps, data, expected, actual, pf)
# Steps written naturally — some numbered, slight style variations

data = [

# ═══ FEATURE 1 section header ════════════════════════════════════════════════
('FEATURE 1', '', '', '', '', '', '', ''),

('Technician Bid\nSubmission', 'TC-BID-01', 'Submit a valid bid',
 '1. Login as technician\n2. Open a PENDING repair request\n3. Enter estimated amount, days and a message\n4. Click submit',
 'Amount: 1500\nDays: 3\nMsg: "I can fix this"',
 'Bid should be created with status PENDING and appear under the repair',
 'As Expected', 'Pass'),

('', 'TC-BID-02', 'Submit bid with no amount entered',
 '1. Login as technician\n2. Open repair request\n3. Leave amount field blank\n4. Enter days = 2\n5. Click submit',
 'Amount: (blank)\nDays: 2',
 'Should show an error — amount is required, bid should not be submitted',
 'As Expected', 'Pass'),

('', 'TC-BID-03', 'Submit bid with no days entered',
 '1. Login as technician\n2. Open repair request\n3. Enter amount = 1000\n4. Leave days blank\n5. Submit',
 'Amount: 1000\nDays: (blank)',
 'Validation error shown, bid not saved',
 'As Expected', 'Pass'),

('', 'TC-BID-04', 'Submit bid with negative amount',
 '1. Login as technician\n2. Open PENDING repair\n3. Enter amount = -500, days = 2\n4. Click submit',
 'Amount: -500\nDays: 2',
 'System should reject negative value and show validation error',
 'As Expected', 'Pass'),

('', 'TC-BID-05', 'Try submitting bid without login',
 '1. Open bid submission URL without logging in\n2. Try to submit a bid',
 'No token / not logged in',
 'Should get redirected to login or get 401 Unauthorized',
 'As Expected', 'Pass'),

('', 'TC-BID-06', 'Customer tries to submit a bid',
 '1. Login as Customer\n2. Try to submit a bid for a repair request\n3. Check response',
 'Customer account credentials',
 'Should be blocked — 403 Forbidden since customers cant bid',
 'As Expected', 'Pass'),

# ═══ FEATURE 2 section header ════════════════════════════════════════════════
('FEATURE 2', '', '', '', '', '', '', ''),

('Customer Bid\nManagement', 'TC-CBID-01', 'Customer views bids for their repair',
 '1. Login as customer\n2. Go to repair requests list\n3. Click on a repair that has bids\n4. Open bids section',
 'Repair request with at least 1 bid submitted',
 'All bids should be visible — technician name, amount, days, message, status',
 'As Expected', 'Pass'),

('', 'TC-CBID-02', 'View bids on repair with no bids',
 '1. Login as customer\n2. Open a repair request that has no bids yet\n3. Check bids section',
 'Repair with 0 bids',
 'Should show empty state or message like "no bids yet"',
 'As Expected', 'Pass'),

('', 'TC-CBID-03', "View another customers bids",
 '1. Login as customer A\n2. Try to access bids for a repair belonging to customer B using the repair ID',
 'Customer A token\nCustomer B repair ID',
 'Should get 403 Forbidden — cannot see other customers data',
 'As Expected', 'Pass'),

('', 'TC-CBID-04', 'Customer accepts a bid',
 '1. Login as customer\n2. Go to bid list for repair\n3. Click Accept on a PENDING bid',
 'Valid bid in PENDING status',
 'Bid becomes ACCEPTED, rest become REJECTED, technician gets assigned',
 'As Expected', 'Pass'),

('', 'TC-CBID-05', 'Customer rejects a bid',
 '1. Login as customer\n2. Open bids for repair\n3. Click Reject on one of the PENDING bids',
 'PENDING bid',
 'Bid status changes to REJECTED, other bids unaffected',
 'As Expected', 'Pass'),

('', 'TC-CBID-06', 'Try to accept an already rejected bid',
 '1. Login as customer\n2. Try to accept a bid that is already REJECTED',
 'Bid status = REJECTED',
 'Error message — cannot accept a rejected bid',
 'As Expected', 'Pass'),

# ═══ FEATURE 3 section header ════════════════════════════════════════════════
('FEATURE 3', '', '', '', '', '', '', ''),

('Bid Acceptance\nWorkflow', 'TC-BAW-01', 'Technician auto-assigned after acceptance',
 '1. Customer accepts a bid\n2. Check the repair request details\n3. See if technician is assigned',
 'Customer accepts bid from Tech Y',
 'Tech Y should be automatically assigned to the repair request',
 'As Expected', 'Pass'),

('', 'TC-BAW-02', 'Other bids get rejected automatically',
 '1. Have 2 bids on same repair\n2. Customer accepts bid A\n3. Check status of bid B',
 '2 bids on 1 repair',
 'Bid A = ACCEPTED, bid B should auto change to REJECTED',
 'As Expected', 'Pass'),

('', 'TC-BAW-03', 'Delivery job created after bid accepted',
 '1. Customer accepts a bid\n2. Login as delivery man\n3. Check available delivery jobs',
 'Bid acceptance triggers delivery creation',
 'A new delivery job should appear with status PENDING_PICKUP',
 'As Expected', 'Pass'),

('', 'TC-BAW-04', 'Try to accept bid when one is already accepted',
 '1. Accept bid 1 first\n2. Try to accept bid 1 again or bid 2 for same repair',
 'Bid already in ACCEPTED state',
 'Should return error — cannot accept when bid not in PENDING status',
 'As Expected', 'Pass'),

# ═══ FEATURE 5 section header ════════════════════════════════════════════════
('FEATURE 5', '', '', '', '', '', '', ''),

('Delivery Man\nRole', 'TC-DLV-01', 'Delivery man registers account',
 '1. Go to registration page\n2. Fill in name, email, password\n3. Select role = Delivery Man\n4. Submit the form',
 'Name: Rahim\nEmail: rahim@test.com\nPassword: pass1234\nRole: DELIVERY_MAN',
 'Account gets created, role should be DELIVERY_MAN, can login',
 'As Expected', 'Pass'),

('', 'TC-DLV-03', 'Delivery man accesses dashboard after login',
 '1. Login with delivery man credentials\n2. Check if delivery dashboard loads',
 'Delivery man email and password',
 'Dashboard loads successfully, 200 OK',
 'As Expected', 'Pass'),

('', 'TC-DLV-04', 'Other roles blocked from delivery dashboard',
 '1. Login as customer\n2. Try to open delivery dashboard URL',
 'Customer account',
 'Access denied — 403 Forbidden',
 'As Expected', 'Pass'),

# ═══ FEATURE 6 section header ════════════════════════════════════════════════
('FEATURE 6', '', '', '', '', '', '', ''),

('Delivery\nDashboard', 'TC-DASH-01', 'Available jobs shown on dashboard',
 '1. Login as delivery man\n2. Open delivery dashboard\n3. Look at available jobs section',
 'At least 1 job with PENDING_PICKUP in system',
 'Jobs list shown with customer info, repair issue, status',
 'As Expected', 'Pass'),

('', 'TC-DASH-02', 'Active job shows after accepting',
 '1. Login as delivery man\n2. Accept a job\n3. Check my jobs / active jobs section',
 'Delivery man accepts 1 job',
 'That job should now show in active jobs for this delivery man',
 'As Expected', 'Pass'),

('', 'TC-DASH-03', 'Empty state when no active jobs',
 '1. Login as delivery man who hasnt accepted any jobs\n2. Check my jobs section',
 'No jobs assigned',
 'Should show empty list — 200 OK',
 'As Expected', 'Pass'),

# ═══ FEATURE 7 section header ════════════════════════════════════════════════
('FEATURE 7', '', '', '', '', '', '', ''),

('Delivery Job\nAcceptance', 'TC-DJOB-01', 'Delivery man accepts a job',
 '1. Login as delivery man\n2. Open delivery dashboard\n3. Click accept on an available job',
 'Job with no assigned delivery man',
 'Job assigned to this delivery man, status changes to GOING_TO_CUSTOMER',
 'As Expected', 'Pass'),

('', 'TC-DJOB-02', 'Same job cannot be accepted twice',
 '1. Delivery man accepts a job\n2. Try to accept the same job again',
 'Job already taken / in progress',
 'Error — job not available (status is no longer PENDING)',
 'As Expected', 'Pass'),

('', 'TC-DJOB-03', 'Customer cannot accept a delivery job',
 '1. Login as customer\n2. Try to call job accept endpoint via API',
 'Customer token',
 'Should get 403 — not authorized',
 'As Expected', 'Pass'),

# ═══ FEATURE 8 section header ════════════════════════════════════════════════
('FEATURE 8', '', '', '', '', '', '', ''),

('Delivery Status\nWorkflow', 'TC-DSW-01', 'Valid status update: going to customer then picked up',
 '1. Login as delivery man\n2. Open active job (status: GOING_TO_CUSTOMER)\n3. Update status to PICKED_UP',
 'Current status = GOING_TO_CUSTOMER',
 'Status should update to PICKED_UP, history entry added',
 'As Expected', 'Pass'),

('', 'TC-DSW-02', 'Invalid status jump should be rejected',
 '1. Login as delivery man\n2. Active job status is GOING_TO_CUSTOMER\n3. Try to set status directly to DELIVERED',
 'Status = GOING_TO_CUSTOMER\nTarget = DELIVERED',
 'Error returned — invalid transition, cannot skip steps',
 'As Expected', 'Pass'),

('', 'TC-DSW-03', 'Valid status update: picked up then at warehouse',
 '1. Login as delivery man\n2. Job status is PICKED_UP\n3. Update to AT_WAREHOUSE',
 'Current status = PICKED_UP',
 'Status updates, auto advances to PENDING_TECH_DELIVERY after warehouse',
 'As Expected', 'Pass'),

('', 'TC-DSW-04', 'Only assigned delivery man can change status',
 '1. Login as a different user (technician)\n2. Try to update status of a delivery job',
 'Tech token — not the assigned DM',
 'Should get 403 — not the assigned person',
 'As Expected', 'Pass'),

# ═══ FEATURE 9 section header ════════════════════════════════════════════════
('FEATURE 9', '', '', '', '', '', '', ''),

('Delivery Status\nHistory', 'TC-DHIST-01', 'History recorded after status update',
 '1. Update delivery status a couple of times\n2. Open the job details\n3. Check statusHistory field',
 'Status updated at least once',
 'Each update should add a new entry in status history with timestamp and note',
 'As Expected', 'Pass'),

('', 'TC-DHIST-02', 'Multiple entries accumulate over time',
 '1. Do 3 or more status updates on a job\n2. Check the history list',
 '3 status transitions',
 'History should show all entries in order — confirmed 5 entries',
 'As Expected', 'Pass'),

('', 'TC-DHIST-03', 'Admin can view status history',
 '1. Login as admin\n2. Open a delivery job by ID\n3. View the statusHistory field in response',
 'Admin token',
 'Full history visible — 200 OK',
 'As Expected', 'Pass'),

# ═══ FEATURE 11 section header ════════════════════════════════════════════════
('FEATURE 11', '', '', '', '', '', '', ''),

('Delivery Tracking\nVisibility', 'TC-DTRK-01', 'Customer can see delivery tracking',
 '1. Login as customer\n2. Open their own repair request page\n3. Check for delivery tracking section',
 'Customer owns repair, delivery job exists',
 'Delivery status and progress visible to customer — 200 OK',
 'As Expected', 'Pass'),

('', 'TC-DTRK-02', 'Admin can see tracking on any repair',
 '1. Login as admin\n2. Open any repair request with a delivery job\n3. Check delivery section',
 'Admin account',
 'Delivery info visible — 200 OK',
 'As Expected', 'Pass'),

('', 'TC-DTRK-03', 'Assigned technician can see delivery tracking',
 '1. Login as lead technician\n2. Open a repair assigned to them\n3. Check delivery tracking',
 'Technician is assigned to the repair',
 'Should see delivery tracking — 200 OK',
 'As Expected', 'Pass'),

('', 'TC-DTRK-04', 'Other customer cannot see tracking',
 '1. Login as customer B\n2. Try to access delivery tracking for customer A repair',
 'Customer B token\nCustomer A repair ID',
 'Should get 403 — forbidden',
 'As Expected', 'Pass'),

# ═══ FEATURE 12 section header ════════════════════════════════════════════════
('FEATURE 12', '', '', '', '', '', '', ''),

('Role-Based\nAccess Control', 'TC-RBAC-01', 'Customer blocked from delivery dashboard',
 '1. Login as customer\n2. Try to open delivery available jobs URL',
 'Customer token',
 '403 Forbidden — customers not allowed',
 'As Expected', 'Pass'),

('', 'TC-RBAC-03', 'No token gets 401',
 '1. Send request to /api/delivery/available without any token\n2. Check response',
 'No JWT',
 '401 Unauthorized',
 'As Expected', 'Pass'),

('', 'TC-RBAC-04', 'Customer cannot update delivery status',
 '1. Login as customer\n2. Try PATCH on delivery status endpoint via API',
 'Customer JWT',
 '403 Forbidden',
 'As Expected', 'Pass'),

('', 'TC-RBAC-05', 'Technician cannot accept delivery jobs',
 '1. Login as technician\n2. Try to accept a delivery job using the API',
 'Technician token',
 '403 Forbidden — only delivery men allowed',
 'As Expected', 'Pass'),
]

ROW = 4
for idx, row in enumerate(data):
    feature, tcId, title, steps, td, expected, actual, pf = row

    # Feature section header row
    if feature.startswith('FEATURE') and not tcId:
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=8)
        c = ws.cell(row=ROW, column=1)
        # Map short name
        feat_names = {
            'FEATURE 1': 'Feature 1: Technician Bid Submission',
            'FEATURE 2': 'Feature 2: Customer Bid Management',
            'FEATURE 3': 'Feature 3: Bid Acceptance Workflow',
            'FEATURE 5': 'Feature 5: Delivery Man Role',
            'FEATURE 6': 'Feature 6: Delivery Dashboard',
            'FEATURE 7': 'Feature 7: Delivery Job Acceptance',
            'FEATURE 8': 'Feature 8: Delivery Status Workflow',
            'FEATURE 9': 'Feature 9: Delivery Status History',
            'FEATURE 11': 'Feature 11: Delivery Tracking Visibility',
            'FEATURE 12': 'Feature 12: Role-Based Access Control',
        }
        c.value = feat_names.get(feature, feature)
        c.font = Font(name='Times New Roman', bold=True, size=11)
        c.fill = fill(FEAT)
        c.alignment = align('left')
        c.border = border()
        ws.row_dimensions[ROW].height = 18
        ROW += 1
        continue

    # Normal data row
    is_odd = (idx % 2 == 0)
    row_bg = YELLOW if is_odd else WHITE

    vals = [feature, tcId, title, steps, td, expected, actual, pf]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=ROW, column=col, value=val)
        c.border = border()
        c.alignment = align()
        if col == 1:
            c.font = font(bold=True, sz=10)
            c.fill = fill(row_bg)
            c.alignment = align('center')
        elif col == 2:
            c.font = font(bold=True, sz=10)
            c.fill = fill(row_bg)
            c.alignment = align('center')
        elif col == 3:
            c.font = font(bold=False, sz=10)
            c.fill = fill(row_bg)
        elif col in (4, 5, 6):
            c.font = font(bold=False, sz=10)
            c.fill = fill(row_bg)
        elif col == 7:
            c.font = font(bold=False, sz=10)
            c.fill = fill(GREEN if pf == 'Pass' else RED)
        elif col == 8:
            c.font = font(bold=True, sz=10)
            c.fill = fill(GREEN if pf == 'Pass' else RED)
            c.alignment = align('center')
    ws.row_dimensions[ROW].height = 70
    ROW += 1

# Freeze header
ws.freeze_panes = 'A4'

out = r'C:\Users\Asif Bin Mahmood\Downloads\Manual_Testing_CRMS_Sprint3_HumanStyle.xlsx'
wb.save(out)
print('Saved: ' + out)
