import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Testing"

thin = Side(style='thin', color='000000')
def B(): return Border(left=thin, right=thin, top=thin, bottom=thin)
def F(c): return PatternFill('solid', fgColor=c)
def AL(h='left', v='center', wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 26
ws.column_dimensions['D'].width = 38
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 36
ws.column_dimensions['G'].width = 20  # Actual Results - YOU fill this
ws.column_dimensions['H'].width = 10  # Pass/Fail - YOU fill this

# Header row
headers = ['Feature', 'Test Case ID', 'Test Title', 'Testing Steps', 'Test Data', 'Expected Results', 'Actual Results', 'Pass/Fail']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(name='Calibri', bold=True, size=11)
    c.fill = F('D9D9D9')
    c.alignment = AL('center')
    c.border = B()
ws.row_dimensions[1].height = 20

# Instruction row
ws.merge_cells('G2:H2')
c = ws['G2']
c.value = '← Fill these two columns after testing each step'
c.font = Font(name='Calibri', italic=True, size=9, color='FF0000')
c.alignment = AL('center')
ws.row_dimensions[2].height = 14

# rows: (feature, tcId, title, steps, data, expected)
# Actual Results and Pass/Fail are LEFT BLANK for you to fill
rows = [
  ('Feature 1 - Technician Bid Submission', 'TC-BID-01', 'Submit a valid bid',
   'Step 1: Login as Technician (lead@crms.test)\nStep 2: Go to Repairs from the sidebar\nStep 3: Open the repair that admin assigned to you\nStep 4: Find the bid form at the bottom of the repair page\nStep 5: Enter Amount = 1500, Days = 3, Message = "I can fix this"\nStep 6: Click Submit Bid',
   'Amount: 1500\nDays: 3\nMessage: I can fix this',
   'Bid is saved. A success message appears. Bid shows up on the repair page with PENDING status.'),

  ('', 'TC-BID-02', 'Bid with no amount',
   'Step 1: Still on the same repair page as technician\nStep 2: Clear the Amount field (leave blank)\nStep 3: Enter Days = 2\nStep 4: Click Submit Bid',
   'Amount: (empty)\nDays: 2',
   'Error message appears. Bid is NOT submitted.'),

  ('', 'TC-BID-03', 'Bid with no days',
   'Step 1: Enter Amount = 1000\nStep 2: Clear the Days field (leave blank)\nStep 3: Click Submit Bid',
   'Amount: 1000\nDays: (empty)',
   'Error message appears. Bid is NOT submitted.'),

  ('', 'TC-BID-04', 'Bid with negative amount',
   'Step 1: Enter Amount = -500\nStep 2: Enter Days = 2\nStep 3: Click Submit Bid',
   'Amount: -500\nDays: 2',
   'Error shown. System rejects negative amount.'),

  ('', 'TC-BID-05', 'Access repair page without login',
   'Step 1: Click Logout\nStep 2: Try to go to http://localhost:5173/repairs directly in the browser',
   'Not logged in',
   'Browser redirects to the login page. Repair page is not accessible.'),

  ('', 'TC-BID-06', 'Customer cannot submit a bid',
   'Step 1: Login as Customer (customer@crms.test / customer123)\nStep 2: Go to My Repairs\nStep 3: Open any repair\nStep 4: Check if a bid submission form is visible',
   'Logged in as Customer',
   'No bid form visible. Customers cannot submit bids.'),

  ('Feature 2 - Customer Bid Management', 'TC-CBID-01', 'Customer views bids',
   'Step 1: Still logged in as Customer\nStep 2: Go to My Repairs from the menu\nStep 3: Open the repair where the technician submitted a bid (TC-BID-01)\nStep 4: Scroll down to the Bids section',
   'Repair has 1 bid from the technician',
   'Bids section shows the bid with technician name, amount, days, and PENDING status.'),

  ('', 'TC-CBID-02', 'Repair with no bids',
   'Step 1: As Customer, open a different repair that has no bids yet\nStep 2: Scroll to the Bids section',
   'Repair with 0 bids',
   'Empty message shown like "No bids yet" or bids section is empty.'),

  ('', 'TC-CBID-03', 'Cannot see another customers repair',
   'Step 1: In the browser address bar, change the repair ID in the URL to a repair that does NOT belong to you\nStep 2: Press Enter',
   'URL with another customers repair ID',
   'Access denied, error page, or redirect. You cannot see another customers repair.'),

  ('', 'TC-CBID-05', 'Customer rejects a bid',
   'Step 1: Open your repair that has a PENDING bid\nStep 2: Click the Reject button next to the bid\nStep 3: Confirm if a dialog appears',
   'Bid is in PENDING status',
   'Bid status changes to REJECTED on screen.'),

  ('', 'TC-CBID-06', 'Try to accept an already rejected bid',
   'Step 1: After rejecting the bid above\nStep 2: Try to click Accept on the same bid (if button still shows)',
   'Bid already REJECTED',
   'Error message shown. Cannot accept a rejected bid.'),

  ('', 'TC-CBID-04', 'Customer accepts a bid',
   'Step 1: Open a repair that has a NEW PENDING bid (ask admin to assign tech and have tech submit another bid)\nStep 2: Click Accept on the bid\nStep 3: Confirm',
   'Bid is in PENDING status',
   'Bid shows ACCEPTED. Other bids show REJECTED. Technician is now assigned.'),

  ('Feature 3 - Bid Acceptance Workflow', 'TC-BAW-01', 'Technician auto-assigned',
   'Step 1: After accepting the bid (TC-CBID-04), stay on the repair detail page\nStep 2: Check the repair info section - look for assigned technician',
   'Bid was just accepted',
   'The technician who bid is now shown as the assigned technician on the repair.'),

  ('', 'TC-BAW-02', 'Other bids auto-rejected',
   'Step 1: On the same repair, scroll to the bids list\nStep 2: Look at the other bids that were not accepted',
   'Multiple bids on same repair',
   'All other bids show REJECTED status automatically.'),

  ('', 'TC-BAW-03', 'Delivery job created',
   'Step 1: Logout\nStep 2: Login as Delivery Man (delivery@crms.test / delivery123)\nStep 3: Open the Delivery Dashboard\nStep 4: Check Available Jobs',
   'Bid was accepted above',
   'A new delivery job appears with PENDING PICKUP status.'),

  ('', 'TC-BAW-04', 'Cannot accept another bid on same repair',
   'Step 1: Login back as Customer\nStep 2: Open the repair that already has an ACCEPTED bid\nStep 3: Try to accept another bid (if any PENDING bids remain)',
   'Repair already has ACCEPTED bid',
   'Error shown. Cannot accept more bids for this repair.'),

  ('Feature 5 - Delivery Man Role', 'TC-DLV-01', 'Delivery man registers',
   'Step 1: Logout completely\nStep 2: Go to http://localhost:5173/register\nStep 3: Fill in Name, Email, Password\nStep 4: Select Role = Delivery Man\nStep 5: Click Register',
   'Name: TestDM\nEmail: testdm@test.com\nPassword: test1234\nRole: Delivery Man',
   'Account created. Can log in with those credentials.'),

  ('', 'TC-DLV-03', 'Delivery man lands on delivery dashboard',
   'Step 1: Login as delivery@crms.test / delivery123\nStep 2: Watch where the page takes you after login',
   'Delivery man credentials',
   'After login, delivery dashboard or delivery section opens automatically.'),

  ('', 'TC-DLV-04', 'Customer cannot access delivery area',
   'Step 1: Login as Customer\nStep 2: Type http://localhost:5173/delivery in the address bar\nStep 3: Press Enter',
   'Logged in as Customer',
   'Access blocked. Page shows error or redirects away.'),

  ('Feature 6 - Delivery Dashboard', 'TC-DASH-01', 'Available jobs shown',
   'Step 1: Login as Delivery Man\nStep 2: Open the Delivery Dashboard\nStep 3: Look at Available Jobs section',
   'At least 1 job with PENDING PICKUP',
   'Job list shows with customer name, repair issue, and status.'),

  ('', 'TC-DASH-02', 'Active job after accepting',
   'Step 1: Accept an available job from the dashboard (click Accept)\nStep 2: Check the My Jobs / Active Jobs section',
   'Delivery man accepts 1 job',
   'That job now appears in active/my jobs section.'),

  ('', 'TC-DASH-03', 'Empty my-jobs at start',
   'Step 1: Login as a fresh Delivery Man who has not accepted anything\nStep 2: Check My Active Jobs section',
   'No jobs assigned yet',
   'Empty list shown or a "no active jobs" message.'),

  ('Feature 7 - Delivery Job Acceptance', 'TC-DJOB-01', 'Accept an available job',
   'Step 1: Login as Delivery Man\nStep 2: Open Delivery Dashboard\nStep 3: Find a job with PENDING PICKUP status\nStep 4: Click Accept Job button',
   'Job with no delivery man assigned',
   'Job is assigned to you. Status changes to GOING TO CUSTOMER on screen.'),

  ('', 'TC-DJOB-02', 'Cannot accept same job twice',
   'Step 1: After accepting a job above\nStep 2: Try to click Accept again on the same job if still visible',
   'Job already accepted',
   'Error shown. Job is no longer in available list.'),

  ('', 'TC-DJOB-03', 'Customer cannot accept delivery jobs',
   'Step 1: Login as Customer\nStep 2: Try to navigate to http://localhost:5173/delivery\nStep 3: Check if any Accept Job button is visible',
   'Logged in as Customer',
   'No accept job button visible. Area is blocked.'),

  ('Feature 8 - Delivery Status Workflow', 'TC-DSW-01', 'Update status: Going to Customer → Picked Up',
   'Step 1: Login as Delivery Man\nStep 2: Open My Active Jobs\nStep 3: Open the active job (status: GOING TO CUSTOMER)\nStep 4: Find status update option\nStep 5: Select PICKED UP and confirm',
   'Current status: GOING TO CUSTOMER',
   'Status updates to PICKED UP on screen.'),

  ('', 'TC-DSW-02', 'Cannot skip status steps',
   'Step 1: On an active job at GOING TO CUSTOMER\nStep 2: Try to select DELIVERED directly from the status dropdown\nStep 3: Click Update',
   'Target: DELIVERED (skipping steps)',
   'Error shown. Invalid transition rejected.'),

  ('', 'TC-DSW-03', 'Update status: Picked Up → At Warehouse',
   'Step 1: On job at PICKED UP status\nStep 2: Select AT WAREHOUSE from status dropdown\nStep 3: Click Update',
   'Current status: PICKED UP',
   'Status updates. May auto-advance to next leg.'),

  ('', 'TC-DSW-04', 'Other role cannot update status',
   'Step 1: Login as Technician or Customer\nStep 2: Try to find any delivery status update button\nStep 3: Try accessing http://localhost:5173/delivery/my-jobs',
   'Logged in as Tech or Customer',
   'No status update option visible. Area is blocked.'),

  ('Feature 9 - Status History', 'TC-DHIST-01', 'History recorded after update',
   'Step 1: Login as Delivery Man\nStep 2: Open an active delivery job\nStep 3: Update the status once\nStep 4: Scroll down to see Status History section',
   'At least 1 status update done',
   'A new history entry is shown with status name and timestamp.'),

  ('', 'TC-DHIST-02', 'Multiple entries accumulate',
   'Step 1: Update status 2 or 3 more times on same job\nStep 2: Check the history list each time',
   '3+ status updates',
   'History list grows with each update. Entries shown in order.'),

  ('', 'TC-DHIST-03', 'Admin sees history',
   'Step 1: Login as Admin (admin@crms.test / admin123)\nStep 2: Open a repair request that has a delivery job\nStep 3: Check the delivery tracking or history section',
   'Admin account',
   'Full status history is visible to admin.'),

  ('Feature 11 - Delivery Tracking', 'TC-DTRK-01', 'Customer sees tracking on their repair',
   'Step 1: Login as Customer\nStep 2: Go to My Repairs\nStep 3: Open the repair that has an active delivery job\nStep 4: Look for delivery tracking section on the page',
   'Customer owns the repair',
   'Delivery tracking section visible showing current status.'),

  ('', 'TC-DTRK-02', 'Admin sees tracking on any repair',
   'Step 1: Login as Admin\nStep 2: Open any repair with a delivery job\nStep 3: Check delivery tracking section',
   'Admin account',
   'Delivery tracking section visible to admin.'),

  ('', 'TC-DTRK-03', 'Technician sees tracking for their repair',
   'Step 1: Login as Lead Technician\nStep 2: Open a repair assigned to you that has delivery\nStep 3: Check delivery tracking section',
   'Technician is assigned to repair',
   'Delivery tracking visible to the assigned technician.'),

  ('', 'TC-DTRK-04', 'Other customer blocked from tracking',
   'Step 1: Login as Customer\nStep 2: In address bar, type the URL of a repair that belongs to a DIFFERENT customer\nStep 3: Try to view the delivery tracking on that page',
   'Different customers repair URL',
   'Access denied or error. Tracking not shown.'),

  ('Feature 12 - Role Based Access', 'TC-RBAC-01', 'Customer blocked from delivery section',
   'Step 1: Login as Customer\nStep 2: In browser address bar go to http://localhost:5173/delivery or /delivery/available\nStep 3: See what happens',
   'Customer account',
   'Access blocked. Redirected or shown an error.'),

  ('', 'TC-RBAC-03', 'Not logged in gets redirected',
   'Step 1: Logout\nStep 2: Try to open http://localhost:5173/repairs or /delivery without logging in',
   'No user logged in',
   'Redirected to login page.'),

  ('', 'TC-RBAC-04', 'Customer cannot update delivery status',
   'Step 1: Login as Customer\nStep 2: Try to find a delivery status update button anywhere in the app\nStep 3: Try http://localhost:5173/delivery',
   'Customer account',
   'No status update button visible. Access blocked.'),

  ('', 'TC-RBAC-05', 'Technician cannot access delivery jobs',
   'Step 1: Login as Lead Technician\nStep 2: Try to go to http://localhost:5173/delivery\nStep 3: Look for any "Accept Job" button',
   'Technician account',
   'Delivery dashboard not accessible. No accept job button.'),
]

ROW_START = 3
for idx, row in enumerate(rows):
    r = ROW_START + idx
    feature, tcId, title, steps, data, expected = row

    vals = [feature, tcId, title, steps, data, expected, '', '']
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.border = B()
        c.font = Font(name='Calibri', size=10)
        c.alignment = AL()

        if col == 1 and val:
            c.font = Font(name='Calibri', bold=True, size=10)
        if col == 2:
            c.alignment = AL('center')

        # Actual Results column - highlighted so you know to fill it
        if col == 7:
            c.fill = F('FFF2CC')  # light orange - fill this yourself
            c.alignment = AL()

        # Pass/Fail column - highlighted
        if col == 8:
            c.fill = F('FFF2CC')  # light orange - fill this yourself
            c.alignment = AL('center')

    ws.row_dimensions[r].height = 90

ws.freeze_panes = 'A3'

out = r'C:\Users\Asif Bin Mahmood\Downloads\Manual_Testing_CRMS_Sprint3_Template.xlsx'
wb.save(out)
print('Done: ' + out)
